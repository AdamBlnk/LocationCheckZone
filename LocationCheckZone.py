from openpyxl import load_workbook
import datetime, os

# compares what is passed to it with the list of HxGN locations
def locationCheck(location, locationList):
    for index in locationList:
        if location == index.strip():
            return True
    
    return False

# opens the path, looks for the 1 .xlsx or .xls file and returns the path to that specific file
def getExcelPath(folderPath):
    fileList = os.listdir(folderPath)
    for file in fileList:
        if file[-5:] == ".xlsx" or file[-4:] == ".xls":
            return str(f"{folderPath}\\{file}")

def getLocationsPath(folderPath):
    fileList = os.listdir(folderPath)
    for file in fileList:
        if file == "locationZone.txt":
            return str(f"{folderPath}\\{file}")



# get the name of the script file
scriptFilename = os.path.abspath(__file__)
# get the folder that contains the script and other files needed
folderPath = os.path.dirname(scriptFilename)

#load the excel file from getExcelPath, returns the only excel file in the folder
wb = load_workbook(getExcelPath(folderPath))

# create variables
mySheet = wb.worksheets[0]
locationFile = open(getLocationsPath(folderPath), "r")
location = ""
locCol = 1
wrongLocations = []

# locationList is just a list of the text file
locationList = locationFile.readlines()

# for each line in the excel sheet, send the location to the method locationCheck
for i in range (2, mySheet.max_row, 1):
    location = mySheet.cell(i, locCol).value

    # If location is not in the list, it will return false
    if locationCheck(location, locationList) == False:
        wrongLocations.append(f"row {i}, {location} is incorrect")

# make a list of all files in the directory
fileList = os.listdir(folderPath)

# if there are any errors, create a .txt of the date and time
if len(wrongLocations) > 0:
    formattedDatetime = datetime.datetime.now().strftime("%m.%d.%Y-%H%M")
    wrongLocationsFile = open(f"{folderPath}\\{formattedDatetime}.txt", "w")    

    # write the errors to console and the .txt
    print("\n")
    for index in wrongLocations:
        wrongLocationsFile.write(f"{index}\n")
        print(index)
    
    input("\nThe results above have been printed to a file named after todays date and time\nPress enter to continue\n")

else:
    print("\nProcess complete\nAll locations accounted for")
    input("Press enter to continue\n")
